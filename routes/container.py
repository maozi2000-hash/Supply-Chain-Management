import os
import zipfile
import io
import json as _json
from datetime import datetime, timezone, timedelta
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, session,
    send_file, jsonify, current_app,
)
from flask_login import login_required
import re
from models import db, Order, BookingRecord, ContainerRecord, ContainerImage, CustomsItem, ActualItem, OrderItem, SkuProduct
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage

container_bp = Blueprint("container", __name__, url_prefix="/container")

DOC_TEMPLATE_PATH = r"C:\Users\actpie\Desktop\整柜报关模板.xlsx"
DOC_EXCHANGE_RATE = 6.9
DOC_CLEARANCE_TOTAL_USD = 10000
DOC_TARGET_PROFIT_RMB = 5000


def _safe_filename(value):
    return re.sub(r'[\\/:*?"<>|]+', "_", value or "N_A")


def _fees_dict(cr):
    """把装柜记录的 5 项费用 + 备注打包成 dict 给前端"""
    return {
        "domestic_transport_fee": float(cr.domestic_transport_fee or 0),
        "ocean_freight_fee":       float(cr.ocean_freight_fee or 0),
        "overseas_truck_fee":      float(cr.overseas_truck_fee or 0),
        "shelving_fee":            float(cr.shelving_fee or 0),
        "other_fee":               float(cr.other_fee or 0),
        "fee_remark":              cr.fee_remark or "",
    }


def _to_fee_float(v):
    """把 Excel 单元格内容转 float，空值视为 0"""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _default_origin_place(order):
    if order and (order.supplier_name or "").strip() == "优瑞奇":
        return "宁波"
    return ""


def _english_destination(destination):
    if not destination:
        return ""
    return destination if re.search(r"[A-Za-z]", destination) else destination


def _allowed_file(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in {"jpg", "jpeg", "png", "gif", "webp", "bmp"}


def _extract_zip_images(container_record, zip_file):
    upload_root = current_app.config["UPLOAD_FOLDER"]
    now = datetime.now()
    month_dir = now.strftime("%Y-%m")
    if container_record.order_id:
        order = db.session.get(Order, container_record.order_id)
        order_no_dir = order.order_no if order else "unknown"
    else:
        order_no_dir = "unknown"
    target_dir = os.path.join(upload_root, month_dir, order_no_dir)
    os.makedirs(target_dir, exist_ok=True)

    saved = []
    try:
        with zipfile.ZipFile(zip_file) as zf:
            for name in zf.namelist():
                if name.endswith("/") or name.startswith("__MACOSX"):
                    continue
                if not _allowed_file(name):
                    continue
                data = zf.read(name)
                try:
                    img = PILImage.open(io.BytesIO(data))
                    img = img.convert("RGB")
                    if img.width > 800:
                        ratio = 800 / img.width
                        img = img.resize((800, int(img.height * ratio)), PILImage.LANCZOS)
                except Exception:
                    continue
                timestamp = now.strftime("%Y%m%d_%H%M%S_%f")
                base = os.path.basename(name) or "image.jpg"
                safe_name = f"装柜照_{timestamp}_{base}"
                abs_path = os.path.join(target_dir, safe_name)
                rel_path = os.path.join(month_dir, order_no_dir, safe_name)
                img.save(abs_path, "JPEG", quality=85)
                saved.append(ContainerImage(
                    container_record_id=container_record.id,
                    file_path=rel_path,
                    original_name=base,
                ))
    except zipfile.BadZipFile:
        return saved, "无效的 ZIP 文件"

    for s in saved:
        db.session.add(s)
    return saved, None


def _save_items(container, customs_json, actual_json):
    # 保护：customs_json/actual_json 不传时（None 或空字符串），保留现有数据不删
    # 仅当明确传了非空 JSON 时才覆盖（避免前端 JS 渲染失败时把数据清空）
    if customs_json is not None and customs_json != "":
        try:
            customs_data = _json.loads(customs_json)
            if isinstance(customs_data, list):
                CustomsItem.query.filter_by(container_record_id=container.id).delete()
                for d in customs_data:
                    sku = d.get("sku", "").strip()
                    qty = d.get("quantity", 0)
                    if sku:
                        db.session.add(CustomsItem(container_record_id=container.id, sku=sku, quantity=qty))
        except (_json.JSONDecodeError, TypeError):
            pass

    if actual_json is not None and actual_json != "":
        try:
            actual_data = _json.loads(actual_json)
            if isinstance(actual_data, list):
                ActualItem.query.filter_by(container_record_id=container.id).delete()
                for d in actual_data:
                    sku = d.get("sku", "").strip()
                    qty = d.get("quantity", 0)
                    if sku:
                        db.session.add(ActualItem(container_record_id=container.id, sku=sku, quantity=qty))
        except (_json.JSONDecodeError, TypeError):
            pass


# ============================================================
@container_bp.route("/")
@login_required
def list_container():
    """统一列表：未装柜的订单 + 已装柜的记录（单表混合）；支持多条件筛选（SKU / 柜号 / 订单号 / 供应商 / 提单号 / 客户名 / 装柜日期）。"""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 15, type=int)
    if per_page not in (20, 50, 100, 200):
        per_page = 15

    sku_query = (request.args.get("sku") or "").strip()
    container_no_query = (request.args.get("container_no") or "").strip()
    order_no_query = (request.args.get("order_no") or "").strip()
    supplier_query = (request.args.get("supplier") or "").strip()
    bl_no_query = (request.args.get("bl_no") or "").strip()
    custom_name_query = (request.args.get("custom_name") or "").strip()
    loading_start = (request.args.get("loading_start") or "").strip()
    loading_end = (request.args.get("loading_end") or "").strip()

    any_filter = any([sku_query, container_no_query, order_no_query,
                      supplier_query, bl_no_query, custom_name_query,
                      loading_start, loading_end])

    pending_subq = db.session.query(ContainerRecord.order_id).distinct()

    pagination = None
    container_records = []
    pending_orders = []
    unified_records = []
    unified_search_results = []
    filter_summary = ""
    matched_filter_labels = []

    if any_filter:
        # ============ 已装柜（ContainerRecord）查询 ============
        join_order = False
        join_booking = False
        c_filters = []
        if container_no_query:
            c_filters.append(ContainerRecord.container_no == container_no_query)
            matched_filter_labels.append(f"柜号={container_no_query}")
        if order_no_query:
            join_order = True
            c_filters.append(Order.order_no == order_no_query)
            matched_filter_labels.append(f"订单号={order_no_query}")
        if supplier_query:
            join_order = True
            c_filters.append(Order.supplier_name.ilike(f"%{supplier_query}%"))
            matched_filter_labels.append(f"供应商含「{supplier_query}」")
        if custom_name_query:
            join_order = True
            c_filters.append(Order.custom_name.ilike(f"%{custom_name_query}%"))
            matched_filter_labels.append(f"客户名含「{custom_name_query}」")
        if bl_no_query:
            join_booking = True
            c_filters.append(BookingRecord.bl_no == bl_no_query)
            matched_filter_labels.append(f"提单号={bl_no_query}")
        if loading_start:
            try:
                d = datetime.strptime(loading_start, "%Y-%m-%d").date()
                c_filters.append(ContainerRecord.loading_date >= d)
                matched_filter_labels.append(f"装柜≥{loading_start}")
            except ValueError:
                loading_start = ""
        if loading_end:
            try:
                d = datetime.strptime(loading_end, "%Y-%m-%d").date()
                c_filters.append(ContainerRecord.loading_date <= d)
                matched_filter_labels.append(f"装柜≤{loading_end}")
            except ValueError:
                loading_end = ""

        cq = ContainerRecord.query
        if join_order:
            cq = cq.join(Order)
        if join_booking:
            cq = cq.join(BookingRecord)
        for f in c_filters:
            cq = cq.filter(f)

        if sku_query:
            sku_cids = [cid for (cid,) in db.session.query(ActualItem.container_record_id)
                         .filter(ActualItem.sku.ilike(f"%{sku_query}%"))
                         .distinct().all()]
            if sku_cids:
                cq = cq.filter(ContainerRecord.id.in_(sku_cids))
            else:
                cq = cq.filter(db.literal(False))
            matched_filter_labels.insert(0, f"SKU 含「{sku_query}」")

        containers = cq.order_by(db.desc(ContainerRecord.id)).distinct().all()

        from collections import OrderedDict
        agg = OrderedDict()
        cids = [cr.id for cr in containers]
        if cids:
            items_q = ActualItem.query.filter(ActualItem.container_record_id.in_(cids))
            if sku_query:
                items_q = items_q.filter(ActualItem.sku.ilike(f"%{sku_query}%"))
            for it in items_q.order_by(ActualItem.id).all():
                agg.setdefault(it.container_record_id, []).append(it)

        for cr in containers:
            items = agg.get(cr.id, [])
            unified_search_results.append({
                "kind": "container",
                "id": cr.id,
                "container": cr,
                "order": cr.order,
                "matched_count": len(items),
                "matched_items": items,
                "fees": _fees_dict(cr),
            })

        # ============ 未装柜（Order）查询 ============
        skip_pending = bool(container_no_query or bl_no_query or loading_start or loading_end)
        if not skip_pending:
            p_filters = []
            if order_no_query:
                p_filters.append(Order.order_no == order_no_query)
            if supplier_query:
                p_filters.append(Order.supplier_name.ilike(f"%{supplier_query}%"))
            if custom_name_query:
                p_filters.append(Order.custom_name.ilike(f"%{custom_name_query}%"))

            pq = Order.query.filter(
                ~Order.id.in_(pending_subq),
                Order.status.notin_(["已取消", "装柜完成"])
            )
            for f in p_filters:
                pq = pq.filter(f)

            if sku_query:
                sku_oids = [oid for (oid,) in db.session.query(OrderItem.order_id)
                             .filter(OrderItem.sku.ilike(f"%{sku_query}%"))
                             .distinct().all()]
                if sku_oids:
                    pq = pq.filter(Order.id.in_(sku_oids))
                else:
                    pq = pq.filter(db.literal(False))

            pending_hits = pq.order_by(db.desc(Order.created_at)).limit(50).all()
            for o in pending_hits:
                matched_items = []
                if sku_query:
                    matched_items = OrderItem.query.filter(
                        OrderItem.order_id == o.id,
                        OrderItem.sku.ilike(f"%{sku_query}%")
                    ).all()
                unified_search_results.append({
                    "kind": "order",
                    "id": o.id,
                    "order": o,
                    "container": None,
                    "matched_count": len(matched_items),
                    "matched_items": matched_items,
                })

        # Sort: pending first (kind order before container), then by -order.id
        unified_search_results.sort(key=lambda r: (0 if r["kind"] == "order" else 1, -r["order"].id))

        filter_summary = " · ".join(matched_filter_labels)
    else:
        # 默认视图：未装柜订单（top 50） + 已装柜订单（分页）合并单表
        pending_orders = Order.query.filter(
            ~Order.id.in_(pending_subq),
            Order.status.notin_(["已取消", "装柜完成"])
        ).order_by(db.desc(Order.created_at)).limit(50).all()
        pagination = ContainerRecord.query.order_by(db.desc(ContainerRecord.id)).paginate(
            page=page, per_page=per_page, error_out=False
        )
        container_records = pagination.items
        # 合并：未装柜（按 created_at desc）+ 已装柜（按 id desc）
        for o in pending_orders:
            unified_records.append({"kind": "order", "id": o.id, "order": o, "container": None, "matched_count": 0, "matched_items": []})
        for cr in container_records:
            unified_records.append({"kind": "container", "id": cr.id, "order": cr.order, "container": cr, "matched_count": 0, "matched_items": [], "fees": _fees_dict(cr)})

    return render_template(
        "container/list.html", active_menu="container",
        unified_records=unified_records,
        unified_search_results=unified_search_results,
        pagination=pagination,
        filter_summary=filter_summary,
        any_filter=any_filter,
        sku_query=sku_query,
        container_no_query=container_no_query,
        order_no_query=order_no_query,
        supplier_query=supplier_query,
        bl_no_query=bl_no_query,
        custom_name_query=custom_name_query,
        loading_start=loading_start,
        loading_end=loading_end,
        per_page=per_page,
    )



@container_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_container():
    view_only = request.args.get("view") == "1"
    container_id = request.args.get("id", type=int)
    orders = Order.query.order_by(Order.created_at.desc()).all()
    bookings = BookingRecord.query.order_by(db.desc(BookingRecord.id)).all()

    target_order_id = request.args.get("order_id") or session.get("last_created_order_id") or ""
    target_order = db.session.get(Order, int(target_order_id)) if target_order_id else None
    selected_order_id = str(target_order.id) if target_order else ""

    # 加载 SKU 体积缓存（传给前端内嵌，避免异步）
    from models import SkuProduct
    all_skus = SkuProduct.query.all()
    sku_volume_map = {}
    sku_weight_map = {}
    sku_cost_map = {}
    for sp in all_skus:
        vol = (sp.length or 0) * (sp.width or 0) * (sp.height or 0) / 1_000_000
        sku_volume_map[sp.sku] = round(vol, 6)
        sku_weight_map[sp.sku] = round(sp.gross_weight or 0, 4)
        sku_cost_map[sp.sku] = round(sp.unit_cost or 0, 4)

    # 对于有 -A/-B/-C 后缀的 SKU，计算同组平均成本
    group_costs = {}  # base_sku -> [costs]
    for sp in all_skus:
        sku = sp.sku or ""
        m = re.match(r"^(.+?)(-[A-Za-z])$", sku)
        if m:
            base = m.group(1)
            if base not in group_costs:
                group_costs[base] = []
            group_costs[base].append(sp.unit_cost or 0)
    for base, costs in group_costs.items():
        if len(costs) > 1:
            avg = sum(costs) / len(costs)
            for sp in all_skus:
                m2 = re.match(r"^(.+?)(-[A-Za-z])$", sp.sku or "")
                if m2 and m2.group(1) == base:
                    sku_cost_map[sp.sku] = round(avg, 4)

    # 编辑模式：加载已有 container
    container = None
    if container_id:
        container = db.session.get(ContainerRecord, container_id)
        if not container:
            flash("装柜记录不存在", "error")
            return redirect(url_for("container.list_container"))
        selected_order_id = str(container.order_id) if container.order_id else ""

    if request.method == "POST":
        container_obj = _build_container_from_form(request, container)
        if container_obj is None:
            customs_data = [{"sku": ci.sku, "quantity": ci.quantity} for ci in container.customs_items] if container else []
            actual_data = [{"sku": ai.sku, "quantity": ai.quantity} for ai in container.actual_items] if container else []
            synced_items = {"customs": customs_data, "actual": actual_data} if container else None
            return render_template(
                "container/form.html", active_menu="container", container=container,
                orders=orders, bookings=bookings, selected_order_id=selected_order_id,
                                        synced_items=synced_items,
                view_only=view_only,
            )

        is_new = container is None
        if is_new:
            existing = ContainerRecord.query.filter_by(container_no=container_obj.container_no).first()
            if existing:
                flash("柜号 " + container_obj.container_no + " 已存在", "warning")
                return redirect(url_for("container.container_detail", id=existing.id))
            db.session.add(container_obj)
            db.session.flush()
        else:
            db.session.flush()

        zip_file = request.files.get("image_zip")
        if zip_file and zip_file.filename:
            saved, err = _extract_zip_images(container_obj, zip_file)
            if err:
                flash(err, "warning")

        customs_json = request.form.get("customs_json")
        actual_json = request.form.get("actual_json")
        # None 时后端不修改现有 items（保护已有数据）
        _save_items(container_obj, customs_json, actual_json)

        order = db.session.get(Order, container_obj.order_id)
        if order and order.status not in ("装柜完成", "已取消"):
            order.status = "装柜完成"

        db.session.commit()
        # 仅创建时清掉 last_created_order_id 标记；更新不清理
        if is_new:
            session.pop("last_created_order_id", None)
        flash("装柜记录已更新" if not is_new else "装柜记录创建成功", "success")
        return redirect(url_for("container.container_detail", id=container_obj.id))

    # GET 请求
    synced_items = None

    # 编辑模式：如果 customs/actual 为空但订单有 items，自动从订单同步（恢复数据）
    if container and (not container.customs_items or not container.actual_items) and container.order and container.order.items:
        from sqlalchemy.orm import joinedload
        order = container.order
        # 用 joinedload 避免 lazy load 失效
        for oi in order.items:
            if not container.customs_items:
                db.session.add(CustomsItem(container_record_id=container.id, sku=oi.sku, quantity=oi.quantity))
            if not container.actual_items:
                db.session.add(ActualItem(container_record_id=container.id, sku=oi.sku, quantity=oi.quantity))
        db.session.commit()
        flash("已自动从订单「" + (order.order_no or "") + "」的明细同步到装柜记录", "info")

    if container:
        customs_data = [{"sku": ci.sku, "quantity": ci.quantity} for ci in container.customs_items]
        actual_data = [{"sku": ai.sku, "quantity": ai.quantity} for ai in container.actual_items]
        synced_items = {"customs": customs_data, "actual": actual_data}
    elif selected_order_id:
        order = db.session.get(Order, int(selected_order_id))
        if order:
            synced_items = [{"sku": oi.sku, "quantity": oi.quantity} for oi in order.items]

    return render_template(
        "container/form.html", active_menu="container", container=container,
        orders=orders, bookings=bookings, selected_order_id=selected_order_id,
        sku_volume_map=sku_volume_map, sku_weight_map=sku_weight_map, sku_cost_map=sku_cost_map,
        synced_items=synced_items,
        target_order=target_order,
        view_only=view_only,
    )


@container_bp.route("/api/container-items/<int:container_id>")
@login_required
def api_container_items(container_id):
    """返回指定柜号所有 actual_items 的 HTML 片段（前端 JS 就地展开调用）
    keyword 参数：只有匹配关键词的 SKU 才加 sku-match 高亮类
    """
    from flask import Response
    cr = db.session.get(ContainerRecord, container_id)
    if not cr:
        return Response("<div class=" + chr(34) + "text-muted p-2" + chr(34) + ">柜号不存在</div>", mimetype="text/html"), 404
    from models import ActualItem
    items = ActualItem.query.filter_by(container_record_id=container_id).order_by(ActualItem.id).all()
    if not items:
        return Response("<div class=" + chr(34) + "text-muted small p-2" + chr(34) + ">柜号 " + cr.container_no + " 暂无装柜明细</div>", mimetype="text/html")
    keyword = (request.args.get("keyword") or "").strip().lower()
    rows = []
    for idx, it in enumerate(items, start=1):
        # 仅当 keyword 非空且 SKU 包含 keyword 时才高亮
        is_match = bool(keyword) and (keyword in it.sku.lower())
        sku_cell = ("<span class=" + chr(34) + "sku-match" + chr(34) + "><strong>" + it.sku + "</strong></span>") if is_match else it.sku
        rows.append(
            "<tr>"
            "<td class=" + chr(34) + "text-muted" + chr(34) + ">" + str(idx) + "</td>"
            "<td>" + sku_cell + "</td>"
            "<td>" + str(it.quantity) + "</td>"
            "</tr>"
        )
    html = (
        '<div class="table-responsive"><table class="table table-sm mb-0">'
        '<thead><tr><th style="width:50px;">序号</th><th>SKU</th><th>数量</th></tr></thead>'
        '<tbody>' + "".join(rows) + '</tbody></table></div>'
    )
    return Response(html, mimetype="text/html")


@container_bp.route("/<int:id>")
@login_required
def container_detail(id):
    container = db.session.get(ContainerRecord, id)
    if not container:
        flash("装柜记录不存在", "error")
        return redirect(url_for("container.list_container"))

    images = sorted(container.images, key=lambda x: x.id, reverse=True)
    customs_items = container.customs_items
    actual_items = container.actual_items
    diff_rows = _compute_diff(customs_items, actual_items)

    # 加载 SKU 产品库，构建 SKU -> 体积/成本映射
    from models import SkuProduct
    sku_products = SkuProduct.query.all()
    sku_map = {}
    for sp in sku_products:
        vol = (sp.length or 0) * (sp.width or 0) * (sp.height or 0) / 1_000_000
        sku_map[sp.sku] = {
            "name": sp.name or "",
            "length": sp.length or 0,
            "width": sp.width or 0,
            "height": sp.height or 0,
            "gross_weight": sp.gross_weight or 0,
            "volume_m3": round(vol, 6),
            "unit_cost": sp.unit_cost or 0,
        }

    # 对于有 -A/-B/-C 后缀的 SKU，计算同组平均成本
    group_costs = {}
    for sku_key, info in sku_map.items():
        m = re.match(r"^(.+?)(-[A-Za-z])$", sku_key)
        if m:
            base = m.group(1)
            if base not in group_costs:
                group_costs[base] = []
            group_costs[base].append(info["unit_cost"])
    for base, costs in group_costs.items():
        if len(costs) > 1:
            avg = sum(costs) / len(costs)
            for sku_key, info in sku_map.items():
                m2 = re.match(r"^(.+?)(-[A-Za-z])$", sku_key)
                if m2 and m2.group(1) == base:
                    info["unit_cost"] = round(avg, 4)

    return render_template(
        "container/detail.html", active_menu="container",
        container=container, images=images,
        customs_items=customs_items, actual_items=actual_items,
        diff_rows=diff_rows, sku_map=sku_map,
    )
@container_bp.route("/<int:id>/delete", methods=["POST"])
@login_required
def delete_container(id):
    container = db.session.get(ContainerRecord, id)
    if not container:
        flash("装柜记录不存在", "error")
        return redirect(url_for("container.list_container"))

    for img in container.images:
        img_abs = os.path.join(current_app.config["UPLOAD_FOLDER"], img.file_path)
        if os.path.exists(img_abs):
            try:
                os.remove(img_abs)
            except OSError:
                pass

    db.session.delete(container)
    db.session.commit()
    flash("装柜记录已删除", "success")
    return redirect(url_for("container.list_container"))


@container_bp.route("/<int:id>/export")
@login_required
def export_container(id):
    container = db.session.get(ContainerRecord, id)
    if not container:
        flash("装柜记录不存在", "error")
        return redirect(url_for("container.list_container"))

    order = container.order
    images = container.images
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        hf = Font(bold=True, size=12, color="FFFFFF")
        hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        ac = Alignment(horizontal="center")
        prefix = order.order_no if order else "N_A"

        def _save_wb(wb, name):
            b = io.BytesIO()
            wb.save(b)
            b.seek(0)
            zf.writestr(name, b.read())

        wb = Workbook()
        ws = wb.active
        ws.title = "订单信息"
        for k, v in [("订单号", order.order_no if order else ""),
                     ("供应商", order.supplier_name if order else ""),
                     ("柜名", order.custom_name if order else ""),
                     ("状态", order.status if order else "")]:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws.cell(row=r, column=2, value=v)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 30
        _save_wb(wb, f"订单信息_{prefix}.xlsx")

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "装柜数据"
        for col, h in enumerate(["柜号","装柜日期","件数","毛重","体积","备注"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ac
        ws2.cell(row=2, column=1, value=container.container_no or "")
        ws2.cell(row=2, column=2, value=container.loading_date.strftime("%Y-%m-%d") if container.loading_date else "")
        ws2.cell(row=2, column=3, value=container.cargo_count or 0)
        ws2.cell(row=2, column=4, value=container.weight or 0)
        ws2.cell(row=2, column=5, value=container.volume or 0)
        ws2.cell(row=2, column=6, value=container.remarks or "")
        for cl in ["A","B","C","D","E","F"]:
            ws2.column_dimensions[cl].width = 18
        _save_wb(wb2, f"装柜数据_{prefix}.xlsx")

        _write_items_sheet(zf, container.customs_items, "报关明细", hf, hfill, ac, order)
        _write_items_sheet(zf, container.actual_items, "真实装柜明细", hf, hfill, ac, order)

        diff_rows = _compute_diff(container.customs_items, container.actual_items)
        wb5 = Workbook()
        ws5 = wb5.active
        ws5.title = "差异明细"
        for col, h in enumerate(["SKU","报关数量","真实数量","差异说明"], 1):
            c = ws5.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ac
        for i, dr in enumerate(diff_rows, start=2):
            ws5.cell(row=i, column=1, value=dr["sku"])
            ws5.cell(row=i, column=2, value=dr["customs_qty"])
            ws5.cell(row=i, column=3, value=dr["actual_qty"])
            ws5.cell(row=i, column=4, value=dr["desc"])
        for cl in ["A","B","C","D"]:
            ws5.column_dimensions[cl].width = 20
        _save_wb(wb5, f"差异明细_{prefix}.xlsx")

        for img in images:
            img_abs = os.path.join(current_app.config["UPLOAD_FOLDER"], img.file_path)
            if os.path.exists(img_abs):
                zf.write(img_abs, f"图片/{img.original_name}")

    zip_buffer.seek(0)
    fn = f"{prefix}_{container.loading_date.strftime('%Y-%m-%d') if container.loading_date else ''}_导出.zip"
    return send_file(zip_buffer, mimetype="application/zip", as_attachment=True, download_name=fn)


@container_bp.route("/<int:id>/export-documents")
@login_required
def export_documents(id):
    container = db.session.get(ContainerRecord, id)
    if not container:
        flash("装柜记录不存在", "error")
        return redirect(url_for("container.list_container"))
    if not os.path.exists(DOC_TEMPLATE_PATH):
        flash("整柜报关模板不存在，请确认桌面模板文件仍在", "error")
        return redirect(url_for("container.container_detail", id=id))

    order = container.order
    booking = container.booking
    today = datetime.now().date()
    contract_no = f"QSYH-{order.order_no if order else container.id}-{today.strftime('%Y%m%d')}"
    sign_date = today - timedelta(days=15)
    ship_date = booking.etd - timedelta(days=2) if booking and booking.etd else None
    destination = booking.destination if booking else ""
    origin_place = _default_origin_place(order)

    wb = load_workbook(DOC_TEMPLATE_PATH)
    ws = wb["报关资料"]

    # 清空并写入商品明细。模板公式覆盖 H:O，最多使用 2:20 共 19 行。
    sku_map = {p.sku: p for p in SkuProduct.query.all()}
    source_items = list(container.customs_items or container.actual_items)
    max_item_rows = 19
    for row in range(2, 21):
        for col in range(1, 7):
            ws.cell(row=row, column=col, value=None)

    for idx, item in enumerate(source_items[:max_item_rows], start=2):
        product = sku_map.get(item.sku)
        volume = 0
        if product:
            volume = (product.length or 0) * (product.width or 0) * (product.height or 0) / 1_000_000
        ws.cell(row=idx, column=1, value=item.sku)
        ws.cell(row=idx, column=2, value=item.quantity or 0)
        ws.cell(row=idx, column=3, value=(product.unit_cost or 0) if product else 0)
        ws.cell(row=idx, column=4, value=volume)
        ws.cell(row=idx, column=5, value=(product.gross_weight or 0) if product else 0)
        ws.cell(row=idx, column=6, value=(product.net_weight or 0) if product else 0)

    ws["X1"] = DOC_EXCHANGE_RATE
    ws["S2"] = contract_no
    ws["V2"] = booking.bl_no if booking else ""
    ws["Y2"] = container.actual_freight or container.estimated_freight or 0
    ws["AA2"] = "宁波" if origin_place == "宁波" else ""
    ws["AC2"] = "NINGBO" if origin_place == "宁波" else ""
    ws["AF2"] = origin_place
    ws["S3"] = sign_date
    ws["V3"] = container.container_no or ""
    ws["Y3"] = "个"
    ws["AA3"] = destination or ""
    ws["AC3"] = _english_destination(destination)
    ws["AF3"] = "美国"
    ws["S4"] = ship_date
    ws["Y4"] = container.weight or 0
    ws["AF4"] = "BY SEA"
    ws["AD7"] = DOC_TARGET_PROFIT_RMB
    ws["AB7"] = f"=({DOC_CLEARANCE_TOTAL_USD}+RANDBETWEEN(1,100))/U7"

    # 暂时不填：出口日期、申报日期、监管方式、征免性质、征免方式、封号。
    # 固定商品资料：无特殊要求时沿用模板默认品名、HS、申报要素等。

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"单证_{_safe_filename(order.order_no if order else str(container.id))}_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@container_bp.route("/<int:id>/export-diff")
@login_required
def export_diff(id):
    container = db.session.get(ContainerRecord, id)
    if not container:
        flash("装柜记录不存在", "error")
        return redirect(url_for("container.list_container"))

    order = container.order
    diff_rows = _compute_diff(container.customs_items, container.actual_items)

    wb = Workbook()
    ws = wb.active
    ws.title = "差异明细"
    hf = Font(bold=True, size=12, color="FFFFFF")
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    for col, h in enumerate(["SKU", "报关数量", "真实数量", "差异说明"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    for i, dr in enumerate(diff_rows, start=2):
        ws.cell(row=i, column=1, value=dr["sku"])
        ws.cell(row=i, column=2, value=dr["customs_qty"])
        ws.cell(row=i, column=3, value=dr["actual_qty"])
        ws.cell(row=i, column=4, value=dr["desc"])
    for cl in ["A","B","C","D"]:
        ws.column_dimensions[cl].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"差异明细_{order.order_no if order else 'N_A'}.xlsx",
    )


@container_bp.route("/order-items/<int:order_id>")
@login_required
def get_order_items(order_id):
    order = db.session.get(Order, order_id)
    if not order:
        return jsonify({"items": [], "message": "订单不存在"})
    items = [{"sku": oi.sku, "quantity": oi.quantity} for oi in order.items]
    return jsonify({"items": items})


@container_bp.route("/import-items", methods=["POST"])
@login_required
def import_items():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择文件"}), 400
    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        return jsonify({"error": "无法读取 Excel 文件"}), 400

    items = []
    for row in rows:
        if not row or not any(row):
            continue
        sku = str(row[0]).strip() if row[0] else ""
        qty_raw = row[1] if len(row) > 1 else 0
        try:
            quantity = int(qty_raw) if qty_raw else 0
        except (ValueError, TypeError):
            continue
        if not sku:
            continue
        items.append({"sku": sku, "quantity": quantity})
    return jsonify({"items": items})


@container_bp.route("/template-items")
@login_required
def template_items():
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    for col, h in enumerate(["SKU", "数量"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="模板.xlsx")


def _build_container_from_form(req, existing):
    order_id = req.form.get("order_id", type=int)
    booking_id = req.form.get("booking_id", type=int) or None
    container_no = req.form.get("container_no", "").strip()
    loading_date = req.form.get("loading_date", "")
    cargo_count = req.form.get("cargo_count", 0, type=int)
    weight = req.form.get("weight", 0, type=float)
    volume = req.form.get("volume", 0, type=float)
    estimated_freight = req.form.get("estimated_freight", 0, type=float)
    actual_freight = req.form.get("actual_freight", 0, type=float)
    remarks = req.form.get("remarks", "").strip()

    if not order_id or not container_no:
        flash("请选择订单并填写柜号", "error")
        return None

    if existing:
        existing.order_id = order_id
        existing.booking_id = booking_id
        existing.container_no = container_no
        existing.loading_date = datetime.strptime(loading_date, "%Y-%m-%d").date() if loading_date else None
        existing.cargo_count = cargo_count
        existing.weight = weight
        existing.volume = volume
        existing.estimated_freight = estimated_freight
        existing.actual_freight = actual_freight
        existing.remarks = remarks
        return existing

    return ContainerRecord(
        order_id=order_id, booking_id=booking_id,
        container_no=container_no,
        loading_date=datetime.strptime(loading_date, "%Y-%m-%d").date() if loading_date else None,
        cargo_count=cargo_count, weight=weight, volume=volume,
        estimated_freight=estimated_freight, actual_freight=actual_freight,
        remarks=remarks,
    )


def _write_items_sheet(zf, items, name, hf, hfill, ac, order):
    wb = Workbook()
    ws = wb.active
    ws.title = name
    for col, h in enumerate(["SKU", "数量"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ac
    for i, it in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=it.sku)
        ws.cell(row=i, column=2, value=it.quantity)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    zf.writestr(f"{name}_{order.order_no if order else 'N_A'}.xlsx", buf.read())


def _compute_diff(customs_items, actual_items):
    customs_map = {}
    for ci in customs_items:
        k = ci.sku.strip().upper()
        customs_map[k] = customs_map.get(k, 0) + ci.quantity

    actual_map = {}
    for ai in actual_items:
        k = ai.sku.strip().upper()
        actual_map[k] = actual_map.get(k, 0) + ai.quantity

    all_keys = set(list(customs_map.keys()) + list(actual_map.keys()))
    result = []
    for key in sorted(all_keys):
        cq = customs_map.get(key, 0)
        aq = actual_map.get(key, 0)
        if cq == aq:
            continue
        orig = key
        for ci in customs_items:
            if ci.sku.strip().upper() == key:
                orig = ci.sku; break
        for ai in actual_items:
            if ai.sku.strip().upper() == key:
                orig = ai.sku; break

        if cq == 0:
            desc = f"仅真实，真实多 {aq}"
        elif aq == 0:
            desc = f"仅报关，报关多 {cq}"
        elif cq > aq:
            desc = f"报关多 {cq - aq}"
        else:
            desc = f"真实多 {aq - cq}"

        result.append({"sku": orig, "customs_qty": cq, "actual_qty": aq, "desc": desc})

    return result


# ============================================================
# 批量导出（按 ids）
# ============================================================
@container_bp.route("/batch-export")
@login_required
def batch_export():
    """支持混合 id：container_ids=1,2 + order_ids=3,4"""
    container_ids_param = (request.args.get("container_ids") or request.args.get("ids") or "").strip()
    order_ids_param = request.args.get("order_ids", "").strip()
    try:
        container_ids = [int(x) for x in container_ids_param.split(",") if x.strip().isdigit()]
    except Exception:
        container_ids = []
    try:
        order_ids = [int(x) for x in order_ids_param.split(",") if x.strip().isdigit()]
    except Exception:
        order_ids = []
    if not container_ids and not order_ids:
        flash("未选择任何记录", "warning")
        return redirect(url_for("container.list_container"))

    records = ContainerRecord.query.filter(ContainerRecord.id.in_(container_ids)).order_by(db.desc(ContainerRecord.id)).all() if container_ids else []
    orders = Order.query.filter(Order.id.in_(order_ids)).order_by(db.desc(Order.id)).all() if order_ids else []
    if not records and not orders:
        flash("所选记录不存在", "warning")
        return redirect(url_for("container.list_container"))

    # 逐个打包为 zip（如有图片则包含图片）
    import zipfile as _zip
    import io as _io
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment

    main_buf = _io.BytesIO()
    with _zip.ZipFile(main_buf, "w", _zip.ZIP_DEFLATED) as zf:
        # 已装柜记录：每个 container 一个 xlsx + 图片
        for cr in records:
            order = cr.order
            prefix = (order.order_no if order else "N_A") + "_" + (cr.container_no or str(cr.id))
            wb = _Workbook()
            ws = wb.active
            ws.title = "装柜数据"
            hf = _Font(bold=True, size=12, color="FFFFFF")
            hfill = _PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            ac = _Alignment(horizontal="center")
            for col, h in enumerate(["柜号", "装柜日期", "件数", "毛重", "体积", "备注"], 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = hf; c.fill = hfill; c.alignment = ac
            ws.cell(row=2, column=1, value=cr.container_no or "")
            ws.cell(row=2, column=2, value=cr.loading_date.strftime("%Y-%m-%d") if cr.loading_date else "")
            ws.cell(row=2, column=3, value=cr.cargo_count or 0)
            ws.cell(row=2, column=4, value=cr.weight or 0)
            ws.cell(row=2, column=5, value=cr.volume or 0)
            ws.cell(row=2, column=6, value=cr.remarks or "")
            for cl in ["A", "B", "C", "D", "E", "F"]:
                ws.column_dimensions[cl].width = 18
            xbuf = _io.BytesIO()
            wb.save(xbuf)
            xbuf.seek(0)
            zf.writestr(f"装柜数据_{prefix}.xlsx", xbuf.read())
            for img in cr.images:
                img_abs = os.path.join(current_app.config["UPLOAD_FOLDER"], img.file_path)
                if os.path.exists(img_abs):
                    zf.write(img_abs, f"图片/{prefix}_{img.original_name}")
        # 未装柜订单：合并成 orders.xlsx
        if orders:
            wb = _Workbook()
            ws = wb.active
            ws.title = "订单"
            hf = _Font(bold=True, size=12, color="FFFFFF")
            hfill = _PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            ac = _Alignment(horizontal="center")
            for col, h in enumerate(["订单号", "供应商", "客户名", "状态", "创建时间"], 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = hf; c.fill = hfill; c.alignment = ac
            for i, o in enumerate(orders, start=2):
                ws.cell(row=i, column=1, value=o.order_no)
                ws.cell(row=i, column=2, value=o.supplier_name)
                ws.cell(row=i, column=3, value=o.custom_name or "")
                ws.cell(row=i, column=4, value=o.status)
                ws.cell(row=i, column=5, value=o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "")
            for cl in ["A", "B", "C", "D", "E"]:
                ws.column_dimensions[cl].width = 20
            obuf = _io.BytesIO()
            wb.save(obuf)
            obuf.seek(0)
            zf.writestr("未装柜订单.xlsx", obuf.read())

    main_buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    # 如果只有 orders，返回 xlsx
    if records and not orders and not orders:
        # Need a different approach: return xlsx instead of zip if only orders
        pass
    if records and not orders:
        # Mixed: return zip
        suffix = "mixed"
    elif orders and not records:
        # Only orders: return xlsx
        from openpyxl import Workbook as __W
        from openpyxl.styles import Font as __F, PatternFill as __P, Alignment as __A
        xb = __W()
        xws = xb.active
        xws.title = "订单"
        for col, h in enumerate(["订单号", "供应商", "客户名", "状态", "创建时间"], 1):
            c = xws.cell(row=1, column=col, value=h)
            c.font = __F(bold=True, size=12, color="FFFFFF")
            c.fill = __P(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            c.alignment = __A(horizontal="center")
        for i, o in enumerate(orders, start=2):
            xws.cell(row=i, column=1, value=o.order_no)
            xws.cell(row=i, column=2, value=o.supplier_name)
            xws.cell(row=i, column=3, value=o.custom_name or "")
            xws.cell(row=i, column=4, value=o.status)
            xws.cell(row=i, column=5, value=o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "")
        for cl in ["A", "B", "C", "D", "E"]:
            xws.column_dimensions[cl].width = 20
        xb_buf = _io.BytesIO()
        xb.save(xb_buf)
        xb_buf.seek(0)
        return send_file(
            xb_buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"订单批量导出_{ts}.xlsx",
        )
    return send_file(
        main_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"装柜批量导出_{ts}.zip",
    )


# ============================================================
# 实际费用管理（人民币）
# ============================================================
FEE_FIELDS = [
    ("domestic_transport_fee", "国内运输费用"),
    ("ocean_freight_fee",       "海运费"),
    ("overseas_truck_fee",      "国外拖车费"),
    ("shelving_fee",            "上架费"),
    ("other_fee",               "其他费用"),
]


@container_bp.route("/<int:id>/save-fees", methods=["POST"])
@login_required
def save_fees(id):
    cr = db.session.get(ContainerRecord, id)
    if not cr:
        return jsonify({"success": False, "error": "装柜记录不存在"}), 404

    cr.domestic_transport_fee = _to_fee_float(request.form.get("domestic_transport_fee"))
    cr.ocean_freight_fee       = _to_fee_float(request.form.get("ocean_freight_fee"))
    cr.overseas_truck_fee      = _to_fee_float(request.form.get("overseas_truck_fee"))
    cr.shelving_fee            = _to_fee_float(request.form.get("shelving_fee"))
    cr.other_fee               = _to_fee_float(request.form.get("other_fee"))
    cr.fee_remark              = (request.form.get("fee_remark") or "").strip()

    db.session.commit()
    total = (cr.domestic_transport_fee + cr.ocean_freight_fee +
             cr.overseas_truck_fee + cr.shelving_fee + cr.other_fee)
    return jsonify({"success": True, "total": round(total, 2)})


@container_bp.route("/template-fees", methods=["GET"])
@login_required
def template_fees():
    """下载实际费用 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "装柜费用"
    headers = ["柜号", "提单号"] + [name for _, name in FEE_FIELDS]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    # 示例行
    sample = ["CAAU1234567", "EGLV143660137161", 1000, 5000, 2000, 500, 300]
    for col, v in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=v)

    # 列宽
    for col_letter in ["A", "B"]:
        ws.column_dimensions[col_letter].width = 22
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 14

    # 说明 sheet
    note = wb.create_sheet("使用说明")
    note["A1"] = "装柜费用导入说明"
    note["A1"].font = Font(bold=True, size=13)
    note["A3"] = "1. 必填一列：柜号 或 提单号 至少填一个"
    note["A4"] = "2. 匹配规则：优先按柜号匹配 → 未匹配则按提单号关联订舱匹配"
    note["A5"] = "3. 5 项费用单位均为人民币（元），留空视为 0（覆盖式）"
    note["A6"] = "4. 提单号填写时：系统会通过 BookingRecord.bl_no 找到对应订舱，再关联装柜记录"
    for r in range(1, 7):
        note.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    note.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="装柜费用导入模板.xlsx",
    )


@container_bp.route("/import-fees", methods=["POST"])
@login_required
def import_fees():
    """Excel 批量导入实际费用，按柜号或提单号匹配
    当携带 container_ids 时（前端勾选行导入），只对选中的 id 生效
    """
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "请选择 .xlsx 文件"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "只支持 .xlsx / .xls 文件"}), 400

    # 解析可选的 container_ids 限制
    selected_ids_raw = (request.form.get("container_ids") or "").strip()
    selected_ids = set()
    selected_only = False
    if selected_ids_raw:
        selected_only = True
        for x in selected_ids_raw.split(","):
            x = x.strip()
            if x.isdigit():
                selected_ids.add(int(x))

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"success": False, "error": f"无法读取 Excel：{e}"}), 400

    success_count = 0
    fail_rows = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        container_no = (str(row[0]).strip() if row and len(row) > 0 and row[0] is not None else "")
        bl_no        = (str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else "")

        if not container_no and not bl_no:
            fail_rows.append({"row": excel_row_idx, "reason": "柜号和提单号都为空"})
            continue

        # 匹配：先柜号，后提单号（提单号通过 BookingRecord 关联）
        cr = None
        match_by = ""
        if container_no:
            cr = ContainerRecord.query.filter_by(container_no=container_no).first()
            if cr:
                match_by = "柜号"
        if not cr and bl_no:
            booking = BookingRecord.query.filter_by(bl_no=bl_no).first()
            if booking and booking.container_records:
                cr = booking.container_records[0]
                match_by = "提单号"
        if not cr:
            fail_rows.append({
                "row": excel_row_idx,
                "reason": f"未找到装柜记录（柜号={container_no or '-'}, 提单号={bl_no or '-'}）"
            })
            continue

        # 勾选模式下，限制只对选中的 id 生效
        if selected_only and cr.id not in selected_ids:
            fail_rows.append({
                "row": excel_row_idx,
                "reason": f"未在勾选列表中（柜号 {cr.container_no}）"
            })
            continue

        # 写入 5 项费用（覆盖式，空值视为 0）
        cr.domestic_transport_fee = _to_fee_float(row[2] if len(row) > 2 else None)
        cr.ocean_freight_fee       = _to_fee_float(row[3] if len(row) > 3 else None)
        cr.overseas_truck_fee      = _to_fee_float(row[4] if len(row) > 4 else None)
        cr.shelving_fee            = _to_fee_float(row[5] if len(row) > 5 else None)
        cr.other_fee               = _to_fee_float(row[6] if len(row) > 6 else None)
        success_count += 1

    if success_count > 0:
        db.session.commit()

    return jsonify({
        "success": True,
        "success_count": success_count,
        "fail_count": len(fail_rows),
        "fail_rows": fail_rows[:50],  # 最多返回 50 条失败明细
    })
