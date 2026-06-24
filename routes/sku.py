import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required
from models import db, SkuProduct
from config import ITEMS_PER_PAGE
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sku_bp = Blueprint("sku", __name__, url_prefix="/sku")


# ============================================================
# 列表
# ============================================================
@sku_bp.route("/")
@login_required
def list_sku():
    page = request.args.get("page", 1, type=int)
<<<<<<< HEAD
    per_page = request.args.get("per_page", ITEMS_PER_PAGE, type=int)
    if per_page not in (20, 50, 100, 200):
        per_page = ITEMS_PER_PAGE
=======
>>>>>>> c26341a738934a24e8a8eb6787eb9988aac4ab69
    keyword = request.args.get("keyword", "").strip()
    sort_by = request.args.get("sort", "sku_group")
    sort_dir = request.args.get("dir", "asc")

    query = SkuProduct.query
    if keyword:
        query = query.filter(
            db.or_(
                SkuProduct.sku.contains(keyword),
                SkuProduct.name.contains(keyword),
            )
        )

    # 产品分组模式：按基础 SKU（去除 -A/-B 等后缀）分组，再按后缀排序
    if sort_by == "sku_group":
        all_results = query.order_by(SkuProduct.sku.asc()).all()
        # 分组排序
        import re as _re
        def sku_sort_key(p):
            sku = p.sku or ""
            # 提取基础 SKU（去除末尾的 -A, -B, -C 等后缀）
            m = _re.match(r"^(.+?)(-[A-Za-z])?$", sku)
            if m:
                base = m.group(1)
                suffix = m.group(2) or ""
                return (base, suffix)
            return (sku, "")

        sorted_results = sorted(all_results, key=sku_sort_key)
        if sort_dir == "desc":
            sorted_results = list(reversed(sorted_results))

        # 手动分页
        total = len(sorted_results)
<<<<<<< HEAD
        start = (page - 1) * per_page
        end = start + per_page
=======
        start = (page - 1) * ITEMS_PER_PAGE
        end = start + ITEMS_PER_PAGE
>>>>>>> c26341a738934a24e8a8eb6787eb9988aac4ab69
        page_items = sorted_results[start:end]

        # 伪造 pagination 对象
        class FakePagination:
            def __init__(self):
                self.items = page_items
                self.page = page
<<<<<<< HEAD
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if total else 1
=======
                self.per_page = ITEMS_PER_PAGE
                self.total = total
                self.pages = (total + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE if total else 1
>>>>>>> c26341a738934a24e8a8eb6787eb9988aac4ab69
            def has_prev(self): return self.page > 1
            def has_next(self): return self.page < self.pages
            @property
            def prev_num(self): return self.page - 1
            @property
            def next_num(self): return self.page + 1
            def iter_pages(self, left_edge=2, left_current=2, right_current=3, right_edge=2):
                last = 0
                for num in range(1, self.pages + 1):
                    if num <= left_edge or (self.page - left_current - 1 < num < self.page + right_current) or num > self.pages - right_edge:
                        if last + 1 != num:
                            yield None
                        yield num
                        last = num

        pagination = FakePagination()
    else:
        sort_columns = {
            "sku": SkuProduct.sku,
            "name": SkuProduct.name,
            "length": SkuProduct.length,
            "width": SkuProduct.width,
            "height": SkuProduct.height,
            "gross_weight": SkuProduct.gross_weight,
            "net_weight": SkuProduct.net_weight,
            "unit_cost": SkuProduct.unit_cost,
            "created_at": SkuProduct.created_at,
        }
        sort_col = sort_columns.get(sort_by, SkuProduct.created_at)
        if sort_dir == "asc":
            query_order = sort_col.asc()
        else:
            query_order = sort_col.desc()
        query = query.order_by(query_order)
<<<<<<< HEAD
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "sku/list.html",
        per_page=per_page,
=======
        pagination = query.paginate(page=page, per_page=ITEMS_PER_PAGE, error_out=False)

    return render_template(
        "sku/list.html",
>>>>>>> c26341a738934a24e8a8eb6787eb9988aac4ab69
        active_menu="sku",
        sku_products=pagination.items,
        keyword=keyword,
        sort_by=sort_by,
        sort_dir=sort_dir,
        pagination=pagination,
    )


# ============================================================
# 新增
# ============================================================
@sku_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_sku():
    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        if not sku:
            flash("SKU 不能为空", "error")
            return render_template("sku/form.html", active_menu="sku", product=None)

        if SkuProduct.query.filter_by(sku=sku).first():
            flash(f"SKU {sku} 已存在", "error")
            return render_template("sku/form.html", active_menu="sku", product=None)

        product = SkuProduct(
            sku=sku,
            name=request.form.get("name", "").strip(),
            length=request.form.get("length", 0, type=float),
            width=request.form.get("width", 0, type=float),
            height=request.form.get("height", 0, type=float),
            gross_weight=request.form.get("gross_weight", 0, type=float),
            net_weight=request.form.get("net_weight", 0, type=float),
            unit_cost=request.form.get("unit_cost", 0, type=float),
            remarks=request.form.get("remarks", "").strip(),
        )
        db.session.add(product)
        db.session.commit()
        flash("SKU 添加成功", "success")
        return redirect(url_for("sku.list_sku"))

    return render_template("sku/form.html", active_menu="sku", product=None)


# ============================================================
# 编辑
# ============================================================
@sku_bp.route("/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_sku(id):
    product = db.session.get(SkuProduct, id)
    if not product:
        flash("SKU 不存在", "error")
        return redirect(url_for("sku.list_sku"))

    if request.method == "POST":
        new_sku_val = request.form.get("sku", "").strip()
        if not new_sku_val:
            flash("SKU 不能为空", "error")
            return render_template("sku/form.html", active_menu="sku", product=product)

        existing = SkuProduct.query.filter_by(sku=new_sku_val).first()
        if existing and existing.id != product.id:
            flash(f"SKU {new_sku_val} 已被使用", "error")
            return render_template("sku/form.html", active_menu="sku", product=product)

        product.sku = new_sku_val
        product.name = request.form.get("name", "").strip()
        product.length = request.form.get("length", 0, type=float)
        product.width = request.form.get("width", 0, type=float)
        product.height = request.form.get("height", 0, type=float)
        product.gross_weight = request.form.get("gross_weight", 0, type=float)
        product.net_weight = request.form.get("net_weight", 0, type=float)
        product.unit_cost = request.form.get("unit_cost", 0, type=float)
        product.remarks = request.form.get("remarks", "").strip()
        db.session.commit()
        flash("SKU 更新成功", "success")
        return redirect(url_for("sku.list_sku"))

    return render_template("sku/form.html", active_menu="sku", product=product)


# ============================================================
# 删除
# ============================================================
@sku_bp.route("/<int:id>/delete", methods=["POST"])
@login_required
def delete_sku(id):
    product = db.session.get(SkuProduct, id)
    if not product:
        flash("SKU 不存在", "error")
        return redirect(url_for("sku.list_sku"))

    db.session.delete(product)
    db.session.commit()
    flash("SKU 已删除", "success")
    return redirect(url_for("sku.list_sku"))


<<<<<<< HEAD
@sku_bp.route("/batch-delete", methods=["POST"])
def batch_delete_sku():
    """批量删除 SKU，需要管理员密码验证（>10个时）"""
    data = request.get_json()
    if not data or "ids" not in data:
        return jsonify({"success": False, "error": "缺少 ids 参数"}), 400

    ids = data.get("ids", [])
    admin_password = data.get("admin_password", "")

    if len(ids) > 10:
        from config import ADMIN_PASSWORD
        if admin_password != ADMIN_PASSWORD:
            return jsonify({"success": False, "error": "管理员密码错误"}), 403

    deleted = 0
    for sid in ids:
        product = db.session.get(SkuProduct, sid)
        if product:
            db.session.delete(product)
            deleted += 1

    db.session.commit()
    return jsonify({"success": True, "deleted": deleted})


=======
>>>>>>> c26341a738934a24e8a8eb6787eb9988aac4ab69
# ============================================================
# AJAX: 按 SKU 查询产品信息（装柜页面调用）
# ============================================================
@sku_bp.route("/lookup")
@login_required
def lookup_sku():
    sku = request.args.get("sku", "").strip()
    if not sku:
        return jsonify(None)
    product = SkuProduct.query.filter_by(sku=sku).first()
    if not product:
        return jsonify(None)
    return jsonify({
        "id": product.id,
        "sku": product.sku,
        "name": product.name or "",
        "length": product.length or 0,
        "width": product.width or 0,
        "height": product.height or 0,
        "gross_weight": product.gross_weight or 0,
        "net_weight": product.net_weight or 0,
        "unit_cost": product.unit_cost or 0,
        # 单件体积 (m3)
        "volume_m3": round((product.length or 0) * (product.width or 0) * (product.height or 0) / 1_000_000, 6),
    })


# ============================================================
# AJAX: 批量查询（装柜页面计算头程单价用）
# ============================================================


# ============================================================
# AJAX: 获取全部 SKU 简要信息（供自动补全下拉使用）
# ============================================================
@sku_bp.route("/all")
@login_required
def all_sku():
    products = SkuProduct.query.order_by(SkuProduct.sku).all()
    result = []
    for p in products:
        result.append({
            "sku": p.sku,
            "name": p.name or "",
            "length": p.length or 0,
            "width": p.width or 0,
            "height": p.height or 0,
            "volume_m3": round((p.length or 0) * (p.width or 0) * (p.height or 0) / 1_000_000, 6),
        })
    return jsonify(result)

@sku_bp.route("/batch-lookup", methods=["POST"])
@login_required
def batch_lookup():
    data = request.get_json(silent=True) or {}
    skus = data.get("skus", [])
    if not skus:
        return jsonify({})
    products = SkuProduct.query.filter(SkuProduct.sku.in_(skus)).all()
    result = {}
    for p in products:
        result[p.sku] = {
            "length": p.length or 0,
            "width": p.width or 0,
            "height": p.height or 0,
            "gross_weight": p.gross_weight or 0,
            "net_weight": p.net_weight or 0,
            "unit_cost": p.unit_cost or 0,
            "volume_m3": round((p.length or 0) * (p.width or 0) * (p.height or 0) / 1_000_000, 6),
        }
    return jsonify(result)


# ============================================================
# 导入 Excel
# ============================================================
@sku_bp.route("/import", methods=["POST"])
@login_required
def import_sku():
    file = request.files.get("file")
    if not file:
        flash("请选择文件", "error")
        return redirect(url_for("sku.list_sku"))

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        flash("无法读取 Excel 文件", "error")
        return redirect(url_for("sku.list_sku"))

    added, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        sku = str(row[0]).strip() if row[0] else ""
        if not sku:
            continue
        if SkuProduct.query.filter_by(sku=sku).first():
            skipped += 1
            continue
        try:
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            length = float(row[2]) if len(row) > 2 and row[2] else 0
            width = float(row[3]) if len(row) > 3 and row[3] else 0
            height = float(row[4]) if len(row) > 4 and row[4] else 0
            gross_weight = float(row[5]) if len(row) > 5 and row[5] else 0
            net_weight = float(row[6]) if len(row) > 6 and row[6] else 0
            unit_cost = float(row[7]) if len(row) > 7 and row[7] else 0
            remarks = str(row[8]).strip() if len(row) > 8 and row[8] else ""
        except (ValueError, TypeError):
            errors.append(f"第{i}行格式错误，已跳过")
            continue

        db.session.add(SkuProduct(
            sku=sku, name=name,
            length=length, width=width, height=height,
            gross_weight=gross_weight, net_weight=net_weight,
            unit_cost=unit_cost, remarks=remarks,
        ))
        added += 1

    db.session.commit()
    msg = f"导入完成：新增 {added} 条"
    if skipped:
        msg += f"，跳过重复 {skipped} 条"
    flash(msg, "success")
    if errors:
        for e in errors[:3]:
            flash(e, "warning")
    return redirect(url_for("sku.list_sku"))


# ============================================================
# 导出 Excel
# ============================================================
@sku_bp.route("/export")
@login_required
def export_sku():
    products = SkuProduct.query.order_by(SkuProduct.sku).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "SKU 产品库"

    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    ac = Alignment(horizontal="center")
    headers = ["SKU", "品名", "长(cm)", "宽(cm)", "高(cm)", "毛重(kg)", "净重(kg)", "采购成本(元)", "单件体积(m3)", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ac

    for i, p in enumerate(products, start=2):
        ws.cell(row=i, column=1, value=p.sku)
        ws.cell(row=i, column=2, value=p.name or "")
        ws.cell(row=i, column=3, value=p.length or 0)
        ws.cell(row=i, column=4, value=p.width or 0)
        ws.cell(row=i, column=5, value=p.height or 0)
        ws.cell(row=i, column=6, value=p.gross_weight or 0)
        ws.cell(row=i, column=7, value=p.net_weight or 0)
        ws.cell(row=i, column=8, value=p.unit_cost or 0)
        vol = round((p.length or 0) * (p.width or 0) * (p.height or 0) / 1_000_000, 6)
        ws.cell(row=i, column=9, value=vol)
        ws.cell(row=i, column=10, value=p.remarks or "")

    for col_letter in ["A","B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="SKU产品库.xlsx",
    )


# ============================================================
# 下载导入模板
# ============================================================
@sku_bp.route("/template")
@login_required
def template_sku():
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU 导入模板"

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    headers = ["SKU", "品名", "长(cm)", "宽(cm)", "高(cm)", "毛重(kg)", "净重(kg)", "采购成本(元)", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="SKU产品库导入模板.xlsx",
    )