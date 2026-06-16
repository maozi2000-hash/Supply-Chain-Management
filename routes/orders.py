import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required
from models import db, Order, OrderItem, ProductionTask, BookingRecord, ContainerRecord
from config import ITEMS_PER_PAGE
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

orders_bp = Blueprint("orders", __name__, url_prefix="/orders")

ORDER_STATUS_OPTIONS = [
    ("下单", "下单"),
    ("生产中", "生产中"),
    ("生产完成", "生产完成"),
    ("订舱中", "订舱中"),
    ("已订舱", "已订舱"),
    ("装柜完成", "装柜完成"),
    ("已取消", "已取消"),
]

ALL_STATUS_VALUES = [v for v, _ in ORDER_STATUS_OPTIONS]


# ============================================================
# 列表
# ============================================================
@orders_bp.route("/")
@login_required
def list_orders():
    page = request.args.get("page", 1, type=int)
    keyword = request.args.get("keyword", "").strip()
    current_status = request.args.get("status", "").strip()

    query = Order.query

    if keyword:
        query = query.filter(
            db.or_(
                Order.order_no.contains(keyword),
                Order.supplier_name.contains(keyword),
            )
        )

    if current_status and current_status in ALL_STATUS_VALUES:
        query = query.filter(Order.status == current_status)

    query = query.order_by(Order.created_at.desc())
    pagination = query.paginate(page=page, per_page=ITEMS_PER_PAGE, error_out=False)

    return render_template(
        "orders/list.html",
        active_menu="orders",
        orders=pagination.items,
        status_options=ORDER_STATUS_OPTIONS,
        current_status=current_status,
        keyword=keyword,
        pagination=pagination,
    )


# ============================================================
# 新增
# ============================================================
@orders_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_order():
    if request.method == "POST":
        order_no = request.form.get("order_no", "").strip()
        supplier_name = request.form.get("supplier_name", "").strip()
        custom_name = request.form.get("custom_name", "").strip()
        status = request.form.get("status", "下单")
        remarks = request.form.get("remarks", "").strip()

        if not order_no or not supplier_name:
            flash("订单号和供应商名称不能为空", "error")
            return render_template(
                "orders/form.html",
                active_menu="orders",
                order=None,
                items=None,
                status_options=ORDER_STATUS_OPTIONS,
            )

        if Order.query.filter_by(order_no=order_no).first():
            flash(f"订单号 {order_no} 已存在", "error")
            return render_template(
                "orders/form.html",
                active_menu="orders",
                order=None,
                items=None,
                status_options=ORDER_STATUS_OPTIONS,
            )

        order = Order(
            order_no=order_no,
            supplier_name=supplier_name,
            custom_name=custom_name,
            status=status,
            remarks=remarks,
        )
        db.session.add(order)
        db.session.flush()

        # 从表单读取明细数据（JSON）
        items_json = request.form.get("items_json", "[]")
        _save_items(order, items_json)

        db.session.commit()
        flash("订单创建成功", "success")
        return redirect(url_for("orders.list_orders"))

    return render_template(
        "orders/form.html",
        active_menu="orders",
        order=None,
        items=None,
        status_options=ORDER_STATUS_OPTIONS,
    )


# ============================================================
# 详情
# ============================================================
@orders_bp.route("/<int:id>")
@login_required
def order_detail(id):
    order = db.session.get(Order, id)
    if not order:
        flash("订单不存在", "error")
        return redirect(url_for("orders.list_orders"))

    items = order.items
    total_quantity = sum(it.quantity for it in items)

    production_tasks = order.production_tasks.order_by(
        db.desc(ProductionTask.id)
    ).all()
    booking_records = order.booking_records.order_by(
        db.desc(BookingRecord.id)
    ).all()
    container_records = order.container_records.order_by(
        db.desc(ContainerRecord.id)
    ).all()

    return render_template(
        "orders/detail.html",
        active_menu="orders",
        order=order,
        items=items,
        total_quantity=total_quantity,
        production_tasks=production_tasks,
        booking_records=booking_records,
        container_records=container_records,
    )


# ============================================================
# 编辑
# ============================================================
@orders_bp.route("/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_order(id):
    order = db.session.get(Order, id)
    if not order:
        flash("订单不存在", "error")
        return redirect(url_for("orders.list_orders"))

    if request.method == "POST":
        order_no = request.form.get("order_no", "").strip()
        supplier_name = request.form.get("supplier_name", "").strip()
        custom_name = request.form.get("custom_name", "").strip()
        status = request.form.get("status", "下单")
        remarks = request.form.get("remarks", "").strip()

        if not order_no or not supplier_name:
            flash("订单号和供应商名称不能为空", "error")
            return render_template(
                "orders/form.html",
                active_menu="orders",
                order=order,
                items=order.items,
                status_options=ORDER_STATUS_OPTIONS,
            )

        existing = Order.query.filter_by(order_no=order_no).first()
        if existing and existing.id != order.id:
            flash(f"订单号 {order_no} 已被使用", "error")
            return render_template(
                "orders/form.html",
                active_menu="orders",
                order=order,
                items=order.items,
                status_options=ORDER_STATUS_OPTIONS,
            )

        order.order_no = order_no
        order.supplier_name = supplier_name
        order.custom_name = custom_name
        order.status = status
        order.remarks = remarks

        # 保存明细
        items_json = request.form.get("items_json", "[]")
        _save_items(order, items_json)

        db.session.commit()
        flash("订单更新成功", "success")
        return redirect(url_for("orders.order_detail", id=order.id))

    items = order.items
    return render_template(
        "orders/form.html",
        active_menu="orders",
        order=order,
        items=items,
        status_options=ORDER_STATUS_OPTIONS,
    )


# ============================================================
# 删除
# ============================================================
@orders_bp.route("/<int:id>/delete", methods=["POST"])
@login_required
def delete_order(id):
    order = db.session.get(Order, id)
    if not order:
        flash("订单不存在", "error")
        return redirect(url_for("orders.list_orders"))

    db.session.delete(order)
    db.session.commit()
    flash("订单已删除", "success")
    return redirect(url_for("orders.list_orders"))


# ============================================================
# 导入 Excel（覆盖替换明细）—— AJAX 接口
# ============================================================
@orders_bp.route("/import-items", methods=["POST"])
@login_required
def import_items():
    """上传 Excel，返回解析后的明细 JSON"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择文件"}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    except Exception:
        return jsonify({"error": "无法读取 Excel 文件，请确认格式正确"}), 400

    items = []
    errors = []
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        sku = str(row[0]).strip() if row[0] else ""
        warehouse = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        qty_raw = row[2] if len(row) > 2 else 0
        try:
            quantity = int(qty_raw) if qty_raw else 0
        except (ValueError, TypeError):
            errors.append(f"第{i}行数量格式错误，已跳过")
            continue

        if not sku:
            continue

        items.append({"sku": sku, "warehouse": warehouse, "quantity": quantity})

    return jsonify({"items": items, "errors": errors})


# ============================================================
# 导出 Excel —— 下载
# ============================================================
@orders_bp.route("/<int:id>/export-items")
@login_required
def export_items(id):
    order = db.session.get(Order, id)
    if not order:
        flash("订单不存在", "error")
        return redirect(url_for("orders.list_orders"))

    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"

    # 表头
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    for col, h in enumerate(["SKU", "仓库", "数量"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, item in enumerate(order.items, start=2):
        ws.cell(row=i, column=1, value=item.sku)
        ws.cell(row=i, column=2, value=item.warehouse or "")
        ws.cell(row=i, column=3, value=item.quantity)

    for col_letter in ["A", "B", "C"]:
        ws.column_dimensions[col_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"订单明细_{order.order_no}.xlsx",
    )


# ============================================================
# 下载导入模板
# ============================================================
@orders_bp.route("/template-items")
@login_required
def template_items():
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    for col, h in enumerate(["SKU", "仓库", "数量"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_letter in ["A", "B", "C"]:
        ws.column_dimensions[col_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="订单明细导入模板.xlsx",
    )


# ============================================================
# 辅助：保存明细
# ============================================================
def _save_items(order, items_json):
    """解析 JSON 格式的明细数据，覆盖替换保存"""
    import json
    try:
        data = json.loads(items_json)
    except (json.JSONDecodeError, TypeError):
        data = []

    # 删除旧明细
    OrderItem.query.filter_by(order_id=order.id).delete()

    # 插入新明细
    for item_data in data:
        sku = item_data.get("sku", "").strip()
        warehouse = item_data.get("warehouse", "").strip()
        quantity = item_data.get("quantity", 0)
        if sku:
            db.session.add(OrderItem(
                order_id=order.id,
                sku=sku,
                warehouse=warehouse,
                quantity=quantity,
            ))
