import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from flask_login import login_required
from models import db, Order, BookingRecord, BookingEtaHistory
from datetime import datetime

booking_bp = Blueprint("booking", __name__, url_prefix="/booking")


VALID_BOOKING_STATUS = ("待订舱", "已订舱")


def _parse_date(value):
    value = (value or "").strip()
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


def _auto_booking_status(*values):
    return "已订舱" if any((v or "").strip() for v in values) else "待订舱"


def _record_eta_history(booking, old_eta, new_eta):
    old_eta = (old_eta or "").strip()
    new_eta = (new_eta or "").strip()
    if old_eta == new_eta:
        return
    db.session.add(BookingEtaHistory(
        booking=booking,
        old_eta=old_eta or None,
        new_eta=new_eta or None,
    ))


@booking_bp.route("/")
@login_required
def list_booking():
    """统一列表：未订舱的订单 + 已订舱的记录（单表混合）；支持多条件筛选。"""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 15, type=int)
    if per_page not in (20, 50, 100, 200):
        per_page = 15

    order_no = (request.args.get("order_no") or "").strip()
    supplier = (request.args.get("supplier") or "").strip()
    bl_no = (request.args.get("bl_no") or "").strip()
    vessel_voyage = (request.args.get("vessel_voyage") or "").strip()
    shipping_company = (request.args.get("shipping_company") or "").strip()
    destination = (request.args.get("destination") or "").strip()
    custom_name = (request.args.get("custom_name") or "").strip()
    current_status = (request.args.get("status") or "").strip()
    if current_status and current_status not in VALID_BOOKING_STATUS:
        current_status = ""
    etd_start = (request.args.get("etd_start") or "").strip()
    etd_end = (request.args.get("etd_end") or "").strip()

    any_filter = any([order_no, supplier, bl_no, vessel_voyage, shipping_company,
                      destination, custom_name, current_status, etd_start, etd_end])

    pending_subq = db.session.query(BookingRecord.order_id).distinct()
    pending_orders = Order.query.filter(
        ~Order.id.in_(pending_subq)
    ).order_by(db.desc(Order.created_at)).limit(50).all()

    booking_records = []
    pagination = None
    unified_records = []
    unified_search_results = []
    filter_summary = ""
    matched_filter_labels = []

    if any_filter:
        join_order = False
        b_filters = []
        if bl_no:
            b_filters.append(BookingRecord.bl_no == bl_no)
            matched_filter_labels.append(f"提单号={bl_no}")
        if vessel_voyage:
            b_filters.append(BookingRecord.vessel_voyage.ilike(f"%{vessel_voyage}%"))
            matched_filter_labels.append(f"船名/航次含「{vessel_voyage}」")
        if shipping_company:
            b_filters.append(BookingRecord.shipping_company.ilike(f"%{shipping_company}%"))
            matched_filter_labels.append(f"船公司含「{shipping_company}」")
        if destination:
            b_filters.append(BookingRecord.destination.ilike(f"%{destination}%"))
            matched_filter_labels.append(f"目的地含「{destination}」")
        if custom_name:
            b_filters.append(BookingRecord.custom_name.ilike(f"%{custom_name}%"))
            matched_filter_labels.append(f"客户名含「{custom_name}」")
        if current_status and current_status in VALID_BOOKING_STATUS:
            matched_filter_labels.append(f"状态={current_status}")
            b_filters.append(BookingRecord.status == current_status)
        if etd_start:
            try:
                d = datetime.strptime(etd_start, "%Y-%m-%d").date()
                b_filters.append(BookingRecord.etd >= d)
                matched_filter_labels.append(f"ETD≥{etd_start}")
            except ValueError:
                etd_start = ""
        if etd_end:
            try:
                d = datetime.strptime(etd_end, "%Y-%m-%d").date()
                b_filters.append(BookingRecord.etd <= d)
                matched_filter_labels.append(f"ETD≤{etd_end}")
            except ValueError:
                etd_end = ""

        if order_no:
            join_order = True
            b_filters.append(Order.order_no == order_no)
            matched_filter_labels.append(f"订单号={order_no}")
        if supplier:
            join_order = True
            b_filters.append(Order.supplier_name.ilike(f"%{supplier}%"))
            matched_filter_labels.append(f"供应商含「{supplier}」")

        bq = BookingRecord.query
        if join_order:
            bq = bq.join(Order)
        for f in b_filters:
            bq = bq.filter(f)
        bq_records = bq.order_by(db.desc(BookingRecord.id)).distinct().all()
        for br in bq_records:
            unified_search_results.append({
                "kind": "booking",
                "id": br.id,
                "booking": br,
                "order": br.order,
                "matched_count": 0,
                "matched_items": [],
            })

        # 未订舱订单：仅适用订单侧字段
        skip_pending = bool(bl_no or vessel_voyage or shipping_company
                            or destination or etd_start or etd_end
                            or (current_status and current_status != "待订舱"))
        if not skip_pending:
            p_filters = []
            if order_no:
                p_filters.append(Order.order_no == order_no)
            if supplier:
                p_filters.append(Order.supplier_name.ilike(f"%{supplier}%"))
            if custom_name:
                p_filters.append(Order.custom_name.ilike(f"%{custom_name}%"))

            pq = Order.query.filter(~Order.id.in_(pending_subq))
            for f in p_filters:
                pq = pq.filter(f)
            pending_hits = pq.order_by(db.desc(Order.created_at)).limit(50).all()
            for o in pending_hits:
                unified_search_results.append({
                    "kind": "order",
                    "id": o.id,
                    "booking": None,
                    "order": o,
                    "matched_count": 0,
                    "matched_items": [],
                })

        # 排序：order 在前，booking 在后
        unified_search_results.sort(key=lambda r: (0 if r["kind"] == "order" else 1, -r["id"]))

        filter_summary = " · ".join(matched_filter_labels)
    else:
        # 默认视图：未订舱（top 50）+ 已订舱（分页）合并单表
        pagination = BookingRecord.query.order_by(db.desc(BookingRecord.id)).paginate(
            page=page, per_page=per_page, error_out=False
        )
        booking_records = pagination.items
        for o in pending_orders:
            unified_records.append({"kind": "order", "id": o.id, "booking": None, "order": o})
        for br in booking_records:
            unified_records.append({"kind": "booking", "id": br.id, "booking": br, "order": br.order})

    return render_template(
        "booking/list.html",
        active_menu="booking",
        unified_records=unified_records,
        unified_search_results=unified_search_results,
        pagination=pagination,
        filter_summary=filter_summary,
        any_filter=any_filter,
        order_no_query=order_no,
        supplier_query=supplier,
        bl_no_query=bl_no,
        vessel_voyage_query=vessel_voyage,
        shipping_company_query=shipping_company,
        destination_query=destination,
        custom_name_query=custom_name,
        current_status=current_status,
        etd_start=etd_start,
        etd_end=etd_end,
        per_page=per_page,
    )



@booking_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_booking():
    view_only = request.args.get("view") == "1"
    pending_subq = db.session.query(BookingRecord.order_id).distinct()
    pending_orders = Order.query.filter(
        ~Order.id.in_(pending_subq)
    ).order_by(db.desc(Order.created_at)).all()
    orders = Order.query.order_by(Order.created_at.desc()).all()
    target_order_id = (
        request.args.get("order_id")
        or session.get("last_created_order_id")
        or (str(pending_orders[0].id) if pending_orders else "")
    )
    target_order = db.session.get(Order, int(target_order_id)) if target_order_id else None
    prefill = {
        "custom_name": target_order.custom_name or "" if target_order else "",
        "vessel_voyage": "", "bl_no": "", "shipping_company": "",
        "etd": "", "eta": "", "destination": "", "cutoff_time": "",
        "remarks": "",
    }
    if request.method == "POST":
        order_id = request.form.get("order_id", type=int)
        custom_name = request.form.get("custom_name", "").strip()
        vessel_voyage = request.form.get("vessel_voyage", "").strip()
        bl_no = request.form.get("bl_no", "").strip()
        shipping_company = request.form.get("shipping_company", "").strip()
        etd = request.form.get("etd", "").strip()
        eta = request.form.get("eta", "").strip()
        destination = request.form.get("destination", "").strip()
        cutoff_time = request.form.get("cutoff_time", "").strip()
        remarks = request.form.get("remarks", "").strip()
        status = _auto_booking_status(vessel_voyage, bl_no, shipping_company, etd, eta, destination, cutoff_time)
        if not order_id:
            flash("请选择关联订单", "error")
            return render_template("booking/form.html", active_menu="booking", booking=None, orders=orders,
                                   target_order=target_order, pending_orders=pending_orders, prefill=prefill,
                                   view_only=view_only)
        br = BookingRecord(
            custom_name=custom_name, vessel_voyage=vessel_voyage, bl_no=bl_no,
            shipping_company=shipping_company,
            etd=_parse_date(etd),
            eta=eta,
            destination=destination, cutoff_time=cutoff_time,
            order_id=order_id, status=status, remarks=remarks,
        )
        db.session.add(br)
        _record_eta_history(br, "", eta)
        order = db.session.get(Order, order_id)
        if order and order.status in ("下单", "生产中", "生产完成", "订舱中"):
            if status == "已订舱":
                order.status = "已订舱"
        db.session.commit()
        session.pop("last_created_order_id", None)
        flash("订舱记录创建成功", "success")
        return redirect(url_for("booking.list_booking"))
    return render_template("booking/form.html", active_menu="booking", booking=None, orders=orders,
                           target_order=target_order, pending_orders=pending_orders, prefill=prefill,
                           view_only=view_only)


@booking_bp.route("/<int:id>")
@login_required
def booking_detail(id):
    booking = db.session.get(BookingRecord, id)
    if not booking:
        flash("订舱记录不存在", "error")
        return redirect(url_for("booking.list_booking"))
    orders = Order.query.order_by(Order.created_at.desc()).all()
    eta_history_rows = BookingEtaHistory.query.filter_by(booking_record_id=id).order_by(
        BookingEtaHistory.changed_at.desc()
    ).all()
    return render_template(
        "booking/form.html",
        active_menu="booking",
        booking=booking,
        orders=orders,
        target_order=booking.order,
        pending_orders=[],
        selected_order_id=str(booking.order_id) if booking.order_id else "",
        view_only=True,
        eta_history_rows=eta_history_rows,
    )


@booking_bp.route("/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_booking(id):
    booking = db.session.get(BookingRecord, id)
    if not booking:
        flash("订舱记录不存在", "error")
        return redirect(url_for("booking.list_booking"))
    orders = Order.query.order_by(Order.created_at.desc()).all()
    if request.method == "POST":
        booking.custom_name = request.form.get("custom_name", "").strip()
        booking.vessel_voyage = request.form.get("vessel_voyage", "").strip()
        booking.bl_no = request.form.get("bl_no", "").strip()
        booking.shipping_company = request.form.get("shipping_company", "").strip()
        etd = request.form.get("etd", "").strip()
        eta = request.form.get("eta", "").strip()
        old_eta = booking.eta or ""
        booking.etd = _parse_date(etd)
        booking.eta = eta
        booking.destination = request.form.get("destination", "").strip()
        booking.cutoff_time = request.form.get("cutoff_time", "").strip()
        posted_order_id = request.form.get("order_id", type=int)
        if posted_order_id:
            booking.order_id = posted_order_id
        booking.remarks = request.form.get("remarks", "").strip()
        booking.status = _auto_booking_status(
            booking.vessel_voyage, booking.bl_no, booking.shipping_company,
            etd, booking.eta, booking.destination, booking.cutoff_time
        )
        _record_eta_history(booking, old_eta, eta)
        order = db.session.get(Order, booking.order_id)
        if order and order.status not in ("装柜完成", "已取消"):
            if booking.status == "已订舱":
                order.status = "已订舱"
        db.session.commit()
        flash("订舱记录更新成功", "success")
        return redirect(url_for("booking.list_booking"))
    return render_template("booking/form.html", active_menu="booking", booking=booking, orders=orders,
                           target_order=booking.order, pending_orders=[], selected_order_id=str(booking.order_id) if booking.order_id else "",
                           view_only=False)


@booking_bp.route("/<int:id>/eta-history")
@login_required
def eta_history(id):
    booking = db.session.get(BookingRecord, id)
    if not booking:
        flash("订舱记录不存在", "error")
        return redirect(url_for("booking.list_booking"))
    history = BookingEtaHistory.query.filter_by(booking_record_id=id).order_by(
        BookingEtaHistory.changed_at.desc()
    ).all()
    return render_template("booking/eta_history.html", active_menu="booking", booking=booking, history=history)


@booking_bp.route("/<int:id>/delete", methods=["POST"])
@login_required
def delete_booking(id):
    br = db.session.get(BookingRecord, id)
    if not br:
        flash("订舱记录不存在", "error")
        return redirect(url_for("booking.list_booking"))
    db.session.delete(br)
    db.session.commit()
    flash("订舱记录已删除", "success")
    return redirect(url_for("booking.list_booking"))


# ============================================================
# 批量导出（按 ids）
# ============================================================
@booking_bp.route("/batch-export")
@login_required
def batch_export():
    """支持混合 id：booking_ids=1,2 + order_ids=3,4"""
    booking_ids_param = (request.args.get("booking_ids") or request.args.get("ids") or "").strip()
    order_ids_param = request.args.get("order_ids", "").strip()
    try:
        booking_ids = [int(x) for x in booking_ids_param.split(",") if x.strip().isdigit()]
    except Exception:
        booking_ids = []
    try:
        order_ids = [int(x) for x in order_ids_param.split(",") if x.strip().isdigit()]
    except Exception:
        order_ids = []
    if not booking_ids and not order_ids:
        flash("未选择任何记录", "warning")
        return redirect(url_for("booking.list_booking"))

    records = BookingRecord.query.filter(BookingRecord.id.in_(booking_ids)).order_by(db.desc(BookingRecord.id)).all() if booking_ids else []
    orders = Order.query.filter(Order.id.in_(order_ids)).order_by(db.desc(Order.id)).all() if order_ids else []
    if not records and not orders:
        flash("所选记录不存在", "warning")
        return redirect(url_for("booking.list_booking"))

    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    if orders and not records:
        # 只有订单：导出订单 xlsx
        from openpyxl import Workbook as __W
        from openpyxl.styles import Font as __F, PatternFill as __P, Alignment as __A
        xb = __W()
        xws = xb.active
        xws.title = "订单"
        for col, h in enumerate(["订单号", "供应商", "客户名", "状态", "创建时间"], 1):
            c = xws.cell(row=1, column=col, value=h)
            c.font = __F(bold=True, size=12, color="FFFFFF")
            c.fill = __P(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            c.alignment = __A(horizontal="center")
        for i, o in enumerate(orders, start=2):
            xws.cell(row=i, column=1, value=o.order_no)
            xws.cell(row=i, column=2, value=o.supplier_name)
            xws.cell(row=i, column=3, value=o.custom_name or "")
            xws.cell(row=i, column=4, value=o.status)
            xws.cell(row=i, column=5, value=o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "")
        for cl in ["A", "B", "C", "D", "E"]:
            xws.column_dimensions[cl].width = 20
        xb_buf = io.BytesIO()
        xb.save(xb_buf)
        xb_buf.seek(0)
        return send_file(
            xb_buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"订单批量导出_{ts}.xlsx",
        )
    # 有订舱记录：导出订舱 xlsx
    buf = io.BytesIO()
    wb = _Workbook()
    ws = wb.active
    ws.title = "订舱记录"
    hf = _Font(bold=True, size=12, color="FFFFFF")
    hfill = _PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    ac = _Alignment(horizontal="center")
    headers = ["订单号", "供应商", "客户名", "船名/航次", "提单号", "船公司", "ETD", "ETA", "目的地", "截单时间", "状态", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ac
    for i, br in enumerate(records, start=2):
        order = br.order
        ws.cell(row=i, column=1, value=order.order_no if order else "")
        ws.cell(row=i, column=2, value=order.supplier_name if order else "")
        ws.cell(row=i, column=3, value=br.custom_name or "")
        ws.cell(row=i, column=4, value=br.vessel_voyage or "")
        ws.cell(row=i, column=5, value=br.bl_no or "")
        ws.cell(row=i, column=6, value=br.shipping_company or "")
        ws.cell(row=i, column=7, value=br.etd.strftime("%Y-%m-%d") if br.etd else "")
        ws.cell(row=i, column=8, value=br.eta or "")
        ws.cell(row=i, column=9, value=br.destination or "")
        ws.cell(row=i, column=10, value=br.cutoff_time or "")
        ws.cell(row=i, column=11, value=br.status or "")
        ws.cell(row=i, column=12, value=br.remarks or "")
    for cl in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[cl].width = 16
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"订舱批量导出_{ts}.xlsx",
    )
